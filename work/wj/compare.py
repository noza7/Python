from openpyxl import Workbook
from openpyxl import load_workbook

path1 = 'byscj_ou专.xlsx'
path2 = '成绩单专.xlsx'

# wb = Workbook(path1)
# ws = wb.loaded_theme(path1)

wb = load_workbook(path2)
ws = wb['sheet1']
ls1=[]
for i in range(1, 11072):
    student_num_name = ws['a{}'.format(i)].value
    student_num = ws['c{}'.format(i)].value
    if student_num_name == '学号：':
        print(student_num)
        ls1.append(student_num)

# print(ls1)